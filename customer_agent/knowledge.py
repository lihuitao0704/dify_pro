"""
增强RAG知识库
数据源 → Knowledge/ 目录：
  公司信息    → 品牌/历程/校区/成功案例
  公司业务    → 留学项目文档（新加坡/德国）
  留学政策    → 德/新 各国 part1-7
  高频问题FAQ → 申请流程/费用/退费
检索算法：关键词重合 + 主题映射 + 子串匹配三重评分 + FAQ精确/模糊匹配
"""

import os
import re
from customer_agent.config import config


# ============================================================
# 文本预处理（复用 student_agent/loader 的算法）
# ============================================================
def _extract_keywords(text: str) -> set:
    """提取中文 2-4 字组 + 英文单词"""
    text_lower = text.lower()
    kws = set()
    # 中文 2-4 gram
    cleans = re.sub(r"[a-zA-Z0-9\s_\-]+", " ", text_lower)
    cleans = cleans.strip()
    for n in (2, 3, 4):
        for i in range(len(cleans) - n + 1):
            kws.add(cleans[i : i + n])
    # 英文单词（3字符以上）
    for m in re.finditer(r"[a-z]{3,}", text_lower):
        kws.add(m.group(0))
    return kws


def _split_chunks(text: str, chunk_size: int = 500, overlap: int = 50) -> list:
    """按段落切分，回退到固定长度"""
    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
    if not paragraphs:
        paragraphs = [text[i : i + chunk_size] for i in range(0, len(text), chunk_size - overlap)]
        return paragraphs

    chunks = []
    buf = ""
    for p in paragraphs:
        if len(buf) + len(p) + 2 <= chunk_size:
            buf += ("\n\n" + p if buf else p)
        else:
            if buf:
                chunks.append(buf)
            buf = p
    if buf:
        chunks.append(buf)
    return chunks


def _load_file(path: str) -> str | None:
    """通用文件加载，按扩展名选解析器"""
    if not os.path.exists(path):
        return None
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in (".txt", ".md", ".text"):
            with open(path, encoding="utf-8") as f:
                return f.read()
        elif ext == ".docx":
            from docx import Document
            doc = Document(path)
            return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
        elif ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            lines = []
            for ws in wb:
                for row in ws.iter_rows(values_only=True):
                    line = " | ".join(str(c) for c in row if c is not None)
                    if line.strip():
                        lines.append(line)
            return "\n".join(lines)
        elif ext == ".pdf":
            import fitz
            doc = fitz.open(path)
            return "\n\n".join(page.get_text() for page in doc)
    except Exception as e:
        print(f"[KB] 加载失败 {path}: {e}")
    return None


def _load_folder(folder: str) -> list:
    """递归加载目录下所有文件 → [{title, content}]"""
    docs = []
    if not os.path.isdir(folder):
        return docs
    for fname in sorted(os.listdir(folder)):
        full = os.path.join(folder, fname)
        if os.path.isdir(full) and not fname.startswith("."):
            docs.extend(_load_folder(full))
        elif not fname.startswith("."):
            content = _load_file(full)
            if content:
                docs.append({"title": fname, "content": content})
    return docs


# ============================================================
# 主题映射：用户口语 → 文档标题关键词
# ============================================================
TOPIC_TITLE_MAP = {
    "公司":      ["公司信息", "品牌", "历程", "成功案例", "校区", "简介"],
    "校区":      ["校区", "分布", "地址"],
    "案例":      ["成功案例", "客户案例", "校友"],
    "德国":      ["德国", "Deutschland", "TU9", "TU", "精英"],
    "新加坡":    ["新加坡", "Singapore", "NUS", "NTU"],
    "政策":      ["留学政策", "签证", "移民", "居留"],
    "签证":      ["签证", "student pass", "d-visa", "续签"],
    "费用":      ["费用", "学费", "收费", "价格", "tuition"],
    "退费":      ["退费", "退款", "退款政策"],
    "流程":      ["流程", "申请步骤", "step", "手续"],
    "申请":      ["申请", "admission", "申请条件", "入学"],
    "语言":      ["语言", "德语", "雅思", "ielts", "托福", "toefl"],
    "院校":      ["大学", "院校", "university", "学校"],
    "深造":      ["硕士", "硕士升读", "本升硕", "读研"],
    "预科":      ["预科", "预科项目"],
    "专升本":    ["专升本"],
    "认证":      ["认证", "中留服"],
}


# ============================================================
# 知识库主体
# ============================================================
class KnowledgeBase:
    """轻量 in-memory RAG，支持 FAQ + 文档检索"""

    def __init__(self):
        self.chunks: list = []          # [{title, text, keywords}]
        self.faq_map: dict = {}         # {question: answer}
        self.faq_keywords: dict = {}    # {question: keyword_set}
        self.doc_count = 0
        self._loaded = False

    def load_all(self):
        """加载全部知识"""
        print("[KB] 开始加载知识库...")
        self._load_company_info()
        self._load_business_docs()
        self._load_policy_docs()
        self._load_faq()
        self._loaded = True
        print(
            f"[KB] 就绪: {len(self.chunks)} chunks, "
            f"{len(self.faq_map)} FAQ, {self.doc_count} docs"
        )

    # ── 公司信息 ─────────────────────────────────────────────
    def _load_company_info(self):
        base = os.path.join(config.KNOWLEDGE_PATH, "公司信息")
        if not os.path.isdir(base):
            print(f"[KB] 跳过公司信息: 目录不存在 {base}")
            return
        for fname in os.listdir(base):
            full = os.path.join(base, fname)
            if os.path.isfile(full) and not fname.startswith("."):
                content = _load_file(full)
                if content:
                    # 尝试解析 Q&A 对（支持 "Q:" 和 "A:" 分隔)
                    self._parse_qa_pairs(content, fname)
                    # 作为全文 chunk 入库
                    self._add_document(fname, content)

    def _parse_qa_pairs(self, content: str, source: str):
        """解析文件中的问答对：支持 Q:/A: 格式 和 Tab 分隔格式"""
        # 方式1: Q:/A: 块
        blocks = re.split(r"\n\s*\n", content)
        count = 0
        for block in blocks:
            block = block.strip()
            if not block:
                continue
            q = ""
            a = ""
            # 尝试 Q: ... A: 模式
            m = re.search(r"Q[:：]\s*(.+?)\s*A[:：]\s*(.+)", block, re.DOTALL)
            if m:
                q, a = m.group(1).strip(), m.group(2).strip()
            else:
                parts = [p.strip() for p in block.split("\n") if p.strip()]
                if len(parts) >= 2 and ("？" in parts[0] or "?" in parts[0]):
                    q, a = parts[0], parts[1]
            if q and a and len(q) < 200:
                self.faq_map[q] = a
                self.faq_keywords[q] = _extract_keywords(q)
                count += 1
        if count:
            print(f"[KB] 公司信息解析问答对: {count} ({source})")

        # 方式2: Tab 分隔（备用）
        for line in content.split("\n"):
            parts = line.split("\t")
            if len(parts) >= 2:
                q, a = parts[0].strip(), parts[1].strip()
                if q and a and len(q) < 200 and ("？" in q or "?" in q):
                    self.faq_map[q] = a
                    self.faq_keywords[q] = _extract_keywords(q)

    # ── 公司业务 ─────────────────────────────────────────────
    def _load_business_docs(self):
        base = os.path.join(config.KNOWLEDGE_PATH, "公司业务")
        if not os.path.isdir(base):
            print(f"[KB] 跳过公司业务: 目录不存在")
            return
        for doc in _load_folder(base):
            self._add_document(doc["title"], doc["content"])

    # ── 留学政策 ─────────────────────────────────────────────
    def _load_policy_docs(self):
        base = os.path.join(config.KNOWLEDGE_PATH, "留学政策")
        if not os.path.isdir(base):
            print(f"[KB] 跳过留学政策: 目录不存在")
            return
        for doc in _load_folder(base):
            self._add_document(doc["title"], doc["content"])

    # ── FAQ ────────────────────────────────────────────────
    def _load_faq(self):
        base = os.path.join(config.KNOWLEDGE_PATH, "高频问题FAQ")
        if not os.path.isdir(base):
            print(f"[KB] 跳过FAQ: 目录不存在")
            return
        for doc in _load_folder(base):
            self._parse_qa_pairs(doc["content"], doc["title"])
            self._add_document(doc["title"], doc["content"])

    # ── 核心：添加文档 ─────────────────────────────────────
    def _add_document(self, title: str, content: str):
        chunks = _split_chunks(content)
        for chunk in chunks:
            self.chunks.append({
                "title": title,
                "text": chunk,
                "keywords": _extract_keywords(chunk),
            })
        self.doc_count += 1

    # ── 检索：文档 RAG ─────────────────────────────────────
    def search(self, query: str, top_k: int = None) -> list:
        """关键词 + 主题映射 + 子串 三重评分"""
        if top_k is None:
            top_k = config.KB_TOP_K
        if not self.chunks:
            return []

        q_kws = _extract_keywords(query)
        # 主题 → 标题关键词
        relevant_titles = set()
        for topic, title_kws in TOPIC_TITLE_MAP.items():
            if topic in query:
                relevant_titles.update(title_kws)

        scored = []
        for chunk in self.chunks:
            score = 0.0
            title_lower = chunk["title"].lower()
            text_lower = chunk["text"].lower()
            # 标题相关性
            for tk in relevant_titles:
                if tk.lower() in title_lower:
                    score += 10
            # 关键词重合
            overlap = len(q_kws & chunk["keywords"])
            score += overlap * 2
            # 子串匹配（兼容短词）
            for qw in q_kws:
                if qw in text_lower:
                    score += 3
                if qw in title_lower:
                    score += 5
            if score > 0:
                scored.append((score, chunk))

        scored.sort(key=lambda x: x[0], reverse=True)
        return [c["text"] for _, c in scored[:top_k]]

    # ── 检索：FAQ ─────────────────────────────────────────
    def faq_match(self, query: str) -> str | None:
        """FAQ精确 → 模糊匹配"""
        # 精确
        q_clean = query.strip()
        if q_clean in self.faq_map:
            return self.faq_map[q_clean]
        # 模糊
        q_kws = _extract_keywords(q_clean)
        best_score = 0
        best_answer = None
        for question, answer in self.faq_map.items():
            qa_kws = self.faq_keywords.get(question, _extract_keywords(question))
            overlap = len(q_kws & qa_kws)
            if overlap > best_score and overlap >= config.FAQ_FUZZY_MIN_OVERLAP:
                best_score = overlap
                best_answer = answer
        return best_answer

    def is_loaded(self) -> bool:
        return self._loaded


# ============================================================
# 全局单例
# ============================================================
_kb: KnowledgeBase | None = None


def get_kb() -> KnowledgeBase:
    """获取知识库单例（懒加载）"""
    global _kb
    if _kb is None:
        _kb = KnowledgeBase()
        _kb.load_all()
    return _kb


def reload_kb():
    """强制重新加载"""
    global _kb
    _kb = KnowledgeBase()
    _kb.load_all()
    return _kb
